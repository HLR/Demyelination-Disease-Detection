#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Feb 10 17:02:29 2019

@author: grollins
"""

# the following is a data scaler which improves performance
from sklearn.preprocessing import StandardScaler 
from sklearn.metrics import confusion_matrix
import experiments
import mlschemes
import features
import numpy as np
from sklearn.cross_validation import train_test_split
from openpyxl import Workbook
from autoencoderRBM import svm_on_BRBM
from sklearn import ensemble


"""
this function creates an array, each representing a cell's data to be plugged
into an Excel sheet.

This function is specialized for the massive printouts below
switches represents whether or not  feature is 'used' for a machine learning
iteration. It's an array of 1 or 0s typically.

rtuple is an array containing scores. Right now, it's an array of increasing
learning. Its first index is the machine learning scheme used
"""
def prepLine(switches, rtuple):
    arrayLine = switches.copy()
    for i in rtuple:
        arrayLine.append(i)
    return arrayLine

"""
This function actually adds a line of information into the excel sheet
line is an array, each item representing a cell in the row to be added
"""
def addLine(line, sheet):
    #rowmax finds the first unoccupied row
    rowmax = sheet.max_row+1
    for i in range(len(line)):
        sheet.cell(column = i + 1, row = rowmax, value = line[i])

"""
configuration variables
"""
trainPack, testPack, numcats = experiments.allDrugs5Cat()
outputFileName = 'f_allDrugs5Cat.xlsx'
f1Threshold = .7
# populate learning Splits with the fractions of the training sets that you
# would like to test. Always include 1.
learningSplits = [1]

#scaling our data seems to improve the SVC's results
scaler = StandardScaler()


"""
confusion matrix work!
"""
"""
print('got here')
clf = ensemble.RandomForestClassifier()
flist = [features.min_value, features.min_value_i, features.max_value]
trainfeatures = features.feature_loader(trainPack[0], flist)
testfeatures = features.feature_loader(testPack[0], flist)

trainfeatures = scaler.fit_transform(trainfeatures)
testfeatures = scaler.transform(testfeatures)

print('got here1')
clf.fit(trainfeatures, trainPack[1])

print('got here2')
y_pred = clf.predict(testfeatures)
print(confusion_matrix(testPack[1], y_pred))
"""


# this initializes our Excel workbook to output
printoutwb = Workbook()
# initializing our Excel sheets to output
ws1 = printoutwb.create_sheet("1 Feature")
ws2 = printoutwb.create_sheet("2 Features")
ws3 = printoutwb.create_sheet("3 Features")
ws4 = printoutwb.create_sheet("4 Features")
ws5 = printoutwb.create_sheet("5 Features")
ws6 = printoutwb.create_sheet("6 Features")
maxws = printoutwb.create_sheet("Highscores")
autows = printoutwb.create_sheet("Autoencoder")
printoutwb.remove(printoutwb['Sheet'])

# this randomly selects 6 of our designed features to test
# this is done because runtime of this file is O(2^n) where n is the 
# number of features tested
randomFeatures = np.random.choice(features.featureList, 6, replace = False)

# this writes the header for every sheet in excel besides the last sheet
for i in range(len(printoutwb.worksheets)-1):
    keyLine = randomFeatures.copy()
    
    #this converts every feature from its ugly computer name to its
    #human-readable name
    for j in range(len(keyLine)):
        keyLine[j] = keyLine[j].__name__
    
    #this converts keyLine from a numpy object to a regular array
    keyLine = keyLine.tolist()
    keyLine.extend(['scheme'])
    for x in learningSplits:
        keyLine.extend([str(x*100) + "% training", 'precision','recall','F1', 'cross-validation mean'])
    addLine(keyLine, printoutwb.worksheets[i])



"""this block will test every possible combination of the loaded features
with every possible machine learning scheme that we load in"""

# this block will turn on or off every feature to use every possible combo
for d0 in [0,1]:
    for d1 in [0,1]:
        for d2 in [0,1]:
            for d3 in [0,1]:
                for d4 in [0,1]:
                    """
                    set above for loop to <<for d4 in [0,1]:>>
                    in final iteration
                    """
                    for d5 in [0,1]:
                        """
                        set above for loop to <<for d5 in [0,1]:>>
                        in final iteration
                        """
                        # dim stores the current combination of features
                        # being used
                        # dim stores the state of each feature
                        dim = [d0,d1,d2,d3,d4,d5]
                        # testList is a list of features being used
                        testList = []
                        
                        # learning curves will store the scores for each
                        # machine learning scheme at each train test split
                        # for this combination of features
                        """
                        plugarrays is of shape
                        [# mlschemes] x [things to print]
                        """
                        plugarrays = []
                        for i in range(len(mlschemes.modelList)):
                            plugarrays.append(dim.copy())
                            plugarrays[i].extend([mlschemes.modelList[i].__name__])
                            
                        
                        # this is where I add to testList
                        if d0 == 1:
                            testList.append(randomFeatures[0])
                        if d1 == 1:
                            testList.append(randomFeatures[1])
                        if d2 == 1:
                            testList.append(randomFeatures[2])
                        if d3 == 1:
                            testList.append(randomFeatures[3])
                        if d4 == 1:
                            testList.append(randomFeatures[4])
                        if d5 == 1:
                            testList.append(randomFeatures[5])
                            
                        # this if statement catches the case where nothing is
                        # tested
                        if not testList:
                            pass
                        else:
                            trainfeatures = features.feature_loader(trainPack[0],\
                                                                    testList)
                            testfeatures = features.feature_loader(testPack[0],\
                                                                   testList)
                            
                            trainfeatures = scaler.fit_transform(trainfeatures)
                            testfeatures = scaler.transform(testfeatures)
                            # BLOCK
                            # this is our learning curve stuff
                            # for each train-test-split
                            for i in learningSplits:
                                trainsplitsamples, _ , trainsplitcats, _ = \
                                train_test_split(trainfeatures, trainPack[1], \
                                                 test_size = i)
                
                                # test every mlscheme on it, and populate
                                # its corresponding location in
                                # learningcurves' array
                                for j in range(len(mlschemes.modelList)):
                                    # j is the row, corresponding to the model
                                    # i is the column, corresponding to the
                                    # test size
                                    score, data, valscores = (\
                                    mlschemes.modelList[j](\
                                            trainsplitsamples,\
                                            trainsplitcats,\
                                            testfeatures,\
                                            testPack[1]
                                            )
                                    )
                                    
                                    plugarrays[j].append(score)
                                    
                                    
                                    plugarrays[j].append(data[0][0])
                                    plugarrays[j].append(data[1][0])
                                    plugarrays[j].append(data[2][0])
                                    
                                    plugarrays[j].append(valscores.mean())
                                    
                                    if data[2][0] > f1Threshold and i == 1:
                                        addLine(plugarrays[j], maxws)
                                    
                                
                                    
                                
                            # BLOCK END FOR LEARNING CURVE IMPLEMENTATION
                            for j in range(len(mlschemes.modelList)):
                                # this is the printout block
                                # add the score to the corresponding feature
                                # sheet. Each sheet number represents a number
                                # of features tested at once
                                if len(testList) == 1:
                                    addLine(plugarrays[j], ws1)
                                elif len(testList) == 2:
                                    addLine(plugarrays[j], ws2)
                                elif len(testList) == 3:
                                    addLine(plugarrays[j], ws3)
                                elif len(testList) == 4:
                                    addLine(plugarrays[j], ws4)
                                elif len(testList) == 5:
                                    addLine(plugarrays[j], ws5)
                                elif len(testList) == 6:
                                    addLine(plugarrays[j], ws6)
                                
                                

# save the file, before it potentially crashes for the autoencoder stuff
printoutwb.save(outputFileName)            

for i in learningSplits:
    plugTuple = []
    trainsplitsamples, _ , trainsplitcats, _ = \
    train_test_split(trainPack[0], trainPack[1], \
                     test_size = i)
    returntuple = svm_on_BRBM(trainsplitsamples, trainsplitcats,\
                              testPack[0], testPack[1])
    plugTuple.append(i)
    plugTuple.append(returntuple[1])
    addLine(plugTuple, autows)
plugTuple = []
plugTuple.append(1)
plugTuple.append(svm_on_BRBM(trainPack[0], trainPack[1], testPack[0], testPack[1])[1])
addLine(plugTuple, autows)

printoutwb.save(outputFileName)